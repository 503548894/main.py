import urllib.request
from lxml import etree
import openpyxl
if __name__ == "__main__":
    mainurl="https://product.cnmo.com/all/product_t1_p{page}.html#allConShaix"
    urls = []   #保存每一个手机的参数地址和概览地址 位置0是概览地址，位置1是参数地址urls里保存元组
    colors= []  #保存手机名、颜色、年份信息
    print("--------开始爬取连接数据---------")
    for i in range(1,166):#166
        print("获取第{page}页数据：".format(page=i))
        url = mainurl.format(page=i)
        rep = urllib.request.urlopen(url)
        htmlstr = str(rep.read(),'gbk');
        html = etree.HTML(htmlstr,etree.HTMLParser())
        href1 = html.xpath('//ul[contains(@class,"all-con-con-ul")]/li/div/a[1]/@href')
        href2 = html.xpath('//ul[contains(@class,"all-con-con-ul")]/li/div/a[2]/@href')
        for j in range(0,len(href1)):
            urls.append((href1[j],href2[j]))
    print("--------开始爬取具体数据---------")
    i=1
    print("共要爬取:"+str(len(urls)))
    for urlchild in urls:
        print("爬取第"+str(i)+"条：")

        try:
            rep2 = urllib.request.urlopen("https:" + urlchild[1])
            htmlstr2 = rep2.read().decode('gbk','ignore')
            html2 = etree.HTML(htmlstr2, etree.HTMLParser())
            title= html2.xpath('//b[contains(@id,"proName")]/a/text()')#proName
            print(title)
            color = html2.xpath('//p[contains(@paramname,"手机颜色")]/@paramvalue')

            rep1 = urllib.request.urlopen("https:" + urlchild[0])
            htmlstr1 = rep1.read().decode('gbk', 'ignore')
            html1 = etree.HTML(htmlstr1, etree.HTMLParser())
            time = html1.xpath('//p[contains(text(),"上市时间")]/text()')
            if len(time) > 0:
                time = time[0].replace('上市时间：', '')
            else:
                time=''

            if len(color) > 0:
                colors.append((title[0],color[0],time))
            else:
                colors.append((title[0],"",time))
        except Exception as e:
            print("爬取页面发生错误:")
            print(e.args)
        i=i+1
    #开始导出excel
    wb = openpyxl.Workbook()
    ws = wb.active
    i=1
    for color in colors:
        ws.cell(row=i,column=1,value=color[0])
        ws.cell(row=i,column=2,value=color[1])
        ws.cell(row=i, column=3, value=color[2])
        i=i+1
    wb.save('export.xlsx')
